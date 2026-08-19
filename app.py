import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from rapidfuzz import process, fuzz
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from fpdf import FPDF
from datetime import datetime


# ============================================================
# PAGE CONFIGURATION
# ============================================================

st.set_page_config(
    page_title="Attendance Data Transformer",
    layout="wide"
)

st.title("Attendance Data Transformer")

st.markdown(
    """
This app transforms your attendance summary data into detailed student
attendance records.

Each student's working days are calculated individually from their own
Present + Absent values, so students who joined late will not incorrectly
receive the maximum working days of the class.
"""
)


# ============================================================
# SESSION STATE
# ============================================================

if 'processed' not in st.session_state:
    st.session_state.processed = False

if 'summary_df' not in st.session_state:
    st.session_state.summary_df = None

if 'detailed_dfs' not in st.session_state:
    st.session_state.detailed_dfs = {}

if 'sorted_class_names' not in st.session_state:
    st.session_state.sorted_class_names = []

if 'working_days' not in st.session_state:
    st.session_state.working_days = None

if 'file_uploader_key' not in st.session_state:
    st.session_state.file_uploader_key = 0

if 'student_working_days' not in st.session_state:
    st.session_state.student_working_days = {}

if 'student_late_days' not in st.session_state:
    st.session_state.student_late_days = {}

if 'student_very_late_days' not in st.session_state:
    st.session_state.student_very_late_days = {}

if 'student_absent_days' not in st.session_state:
    st.session_state.student_absent_days = {}


# ============================================================
# RESET APPLICATION
# ============================================================

def reset_application():

    st.session_state.processed = False
    st.session_state.summary_df = None
    st.session_state.detailed_dfs = {}
    st.session_state.sorted_class_names = []
    st.session_state.working_days = None

    st.session_state.file_uploader_key += 1

    st.session_state.student_working_days = {}
    st.session_state.student_late_days = {}
    st.session_state.student_very_late_days = {}
    st.session_state.student_absent_days = {}


# ============================================================
# EXCEL STYLING
# ============================================================

def apply_excel_styling(
    worksheet,
    title,
    is_summary=False,
    student_names=None,
    late_threshold=0,
    very_late_threshold=0,
    absent_threshold=0
):

    header_font = Font(
        name='Aptos Display',
        size=12,
        bold=True
    )

    data_font = Font(
        name='Aptos Display',
        size=12
    )

    header_fill = PatternFill(
        start_color="DDEBF7",
        end_color="DDEBF7",
        fill_type="solid"
    )

    yellow_fill = PatternFill(
        start_color="FFFF00",
        end_color="FFFF00",
        fill_type="solid"
    )

    red_fill = PatternFill(
        start_color="F94949",
        end_color="F94949",
        fill_type="solid"
    )

    alignment_center = Alignment(
        horizontal='center',
        vertical='center'
    )

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # --------------------------------------------------------
    # HEADER STYLE
    # --------------------------------------------------------

    for cell in worksheet[1]:

        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = thin_border

    # --------------------------------------------------------
    # DATA STYLE
    # --------------------------------------------------------

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row
    ):

        late_value = 0
        very_late_value = 0
        absent_value = 0

        try:

            # Columns:
            # 0 = Admission No
            # 1 = Student Name
            # 2 = Working Days
            # 3 = Present
            # 4 = Absent
            # 5 = Late
            # 6 = Very Late
            # 7 = Attendance %
            
            late_value = row[5].value or 0
            very_late_value = row[6].value or 0
            absent_value = row[4].value or 0

        except Exception:
            pass

        try:
            late_value = float(late_value)
        except Exception:
            late_value = 0

        try:
            very_late_value = float(very_late_value)
        except Exception:
            very_late_value = 0

        try:
            absent_value = float(absent_value)
        except Exception:
            absent_value = 0

        is_absent = absent_value >= absent_threshold

        for idx, cell in enumerate(row):

            cell.font = data_font
            cell.border = thin_border
            cell.alignment = alignment_center

            if not is_summary:

                # ------------------------------------------------
                # ABSENT
                # ------------------------------------------------

                if is_absent and idx in [1, 4]:

                    cell.fill = red_fill

                # ------------------------------------------------
                # LATE
                # ------------------------------------------------

                elif (
                    late_value >= late_threshold
                    and idx == 5
                ):

                    cell.fill = yellow_fill

                # ------------------------------------------------
                # VERY LATE
                # ------------------------------------------------

                elif (
                    very_late_value >= very_late_threshold
                    and idx == 6
                ):

                    cell.fill = yellow_fill

    # --------------------------------------------------------
    # COLUMN WIDTHS
    # --------------------------------------------------------

    if is_summary:

        column_widths = {
            'A': 20,
            'B': 18,
            'C': 22,
            'D': 15,
            'E': 15,
            'F': 15,
            'G': 18,
            'H': 28
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        worksheet.freeze_panes = "A3"

        # ----------------------------------------------------
        # FIND TOP PERFORMING CLASS
        # ----------------------------------------------------

        top_attendance = -1
        top_row_index = None

        for idx, row in enumerate(
            worksheet.iter_rows(
                min_row=3,
                max_row=worksheet.max_row
            ),
            start=3
        ):

            try:

                attendance = float(row[7].value)

                if attendance > top_attendance:

                    top_attendance = attendance
                    top_row_index = idx

            except Exception:
                pass

        # ----------------------------------------------------
        # DASHBOARD ROW STYLING
        # ----------------------------------------------------

        for i, row in enumerate(
            worksheet.iter_rows(
                min_row=3,
                max_row=worksheet.max_row
            ),
            start=3
        ):

            base_fill = PatternFill(
                start_color="F7F9FC" if i % 2 == 0 else "FFFFFF",
                end_color="F7F9FC" if i % 2 == 0 else "FFFFFF",
                fill_type="solid"
            )

            if i == top_row_index:

                fill = PatternFill(
                    start_color="60D276",
                    end_color="60D276",
                    fill_type="solid"
                )

            else:

                fill = base_fill

            for cell in row:

                cell.fill = fill

                cell.font = Font(
                    name='Aptos Display',
                    size=11
                )

                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center'
                )

                cell.border = thin_border

    else:

        column_widths = {
            'A': 15,
            'B': 40,
            'C': 15,
            'D': 10,
            'E': 10,
            'F': 10,
            'G': 12,
            'H': 15,
            'I': 14
        }

    for col, width in column_widths.items():

        worksheet.column_dimensions[col].width = width

    # --------------------------------------------------------
    # FORMAT ATTENDANCE PERCENTAGE
    # --------------------------------------------------------

    for row in range(
        2,
        worksheet.max_row + 1
    ):

        worksheet[f'H{row}'].number_format = '0.00'

    # --------------------------------------------------------
    # TITLE ROW
    # --------------------------------------------------------

    worksheet.insert_rows(1)

    if is_summary:
        worksheet.merge_cells('A1:H1')
    else:
        worksheet.merge_cells('A1:I1')

    title_cell = worksheet['A1']

    title_cell.value = title

    if is_summary:

        title_cell.font = Font(
            name='Aptos Display',
            size=36,
            bold=True
        )

        worksheet.row_dimensions[1].height = 45

    else:

        title_cell.font = Font(
            name='Aptos Display',
            size=24,
            bold=True
        )

    title_cell.alignment = Alignment(
        horizontal='center',
        vertical='center'
    )

    return worksheet


# ============================================================
# DETECT WORKING DAYS
# ============================================================

def detect_working_days(df):

    """
    This function ONLY detects the maximum working days found
    in the uploaded file for informational purposes.

    It does NOT assign this value to every student.

    Each student's actual working days are calculated separately
    inside process_real_data():

        Working Days = Present + Absent
    """

    try:

        present_col = find_best_column(
            df,
            'Present',
            None
        )

        absent_col = find_best_column(
            df,
            'Absent',
            None
        )

        if present_col and absent_col:

            present = pd.to_numeric(
                df[present_col],
                errors='coerce'
            ).fillna(0)

            absent = pd.to_numeric(
                df[absent_col],
                errors='coerce'
            ).fillna(0)

            student_working_days = present + absent

            if not student_working_days.empty:

                return int(
                    student_working_days.max()
                )

    except Exception:
        pass

    return None


# ============================================================
# FIND BEST COLUMN
# ============================================================

def find_best_column(
    df,
    target_name,
    fallback=None
):

    available_columns = df.columns.tolist()

    # Exact match
    if target_name in available_columns:
        return target_name

    # Case insensitive
    for col in available_columns:

        if (
            str(col).strip().lower()
            == target_name.strip().lower()
        ):

            return col

    # Fuzzy match
    match = process.extractOne(
        target_name,
        available_columns,
        scorer=fuzz.token_sort_ratio
    )

    if match and match[1] > 60:

        return match[0]

    return fallback


# ============================================================
# EXTRACT DATE RANGE
# ============================================================

def extract_date_range(filename):

    """
    Example:

    Attendance Summary_2026-05-04_2026-05-08_20260512113609.xlsx
    """

    try:

        matches = re.findall(
            r'(\d{4}-\d{2}-\d{2})',
            filename
        )

        if len(matches) >= 2:

            start_date = datetime.strptime(
                matches[0],
                "%Y-%m-%d"
            )

            end_date = datetime.strptime(
                matches[1],
                "%Y-%m-%d"
            )

            return (
                start_date.strftime("%d.%m.%y"),
                end_date.strftime("%d.%m.%y")
            )

    except Exception:
        pass

    return ("N/A", "N/A")


# ============================================================
# PDF REPORT
# ============================================================

def generate_pdf_report(
    summary_df,
    detailed_dfs,
    sorted_class_names,
    uploaded_filename
):

    start_date, end_date = extract_date_range(
        uploaded_filename
    )

    pdf = FPDF()

    pdf.set_auto_page_break(
        auto=True,
        margin=10
    )

    # ========================================================
    # SUMMARY PAGE
    # ========================================================

    pdf.add_page()

    pdf.set_font(
        "Arial",
        'B',
        16
    )

    pdf.cell(
        0,
        10,
        f"ATTENDANCE SUMMARY - {start_date} - {end_date}",
        ln=True,
        align='C'
    )

    pdf.ln(5)

    headers = [
        "Class",
        "Students",
        "Days",
        "Present",
        "Absent",
        "Late",
        "V.Late",
        "Attendance %"
    ]

    col_widths = [
        35,
        20,
        18,
        22,
        22,
        18,
        20,
        30
    ]

    pdf.set_font(
        "Arial",
        'B',
        10
    )

    for i, header in enumerate(headers):

        pdf.cell(
            col_widths[i],
            10,
            header,
            border=1,
            align='C'
        )

    pdf.ln()

    pdf.set_font(
        "Arial",
        '',
        9
    )

    for _, row in summary_df.iterrows():

        values = [
            str(row["Class"]),
            str(row["Total_Students"]),
            str(row["Total_Working_Days"]),
            str(row["Avg_Present"]),
            str(row["Avg_Absent"]),
            str(row["Avg_Late"]),
            str(row["Avg_Very_Late"]),
            f'{row["Avg_Attendance_Percentage"]:.2f}'
        ]

        for i, value in enumerate(values):

            pdf.cell(
                col_widths[i],
                9,
                value,
                border=1,
                align='C'
            )

        pdf.ln()

    # ========================================================
    # DETAILED CLASS PAGES
    # ========================================================

    for class_name in sorted_class_names:

        if class_name not in detailed_dfs:
            continue

        df_detail = detailed_dfs[class_name]

        pdf.add_page()

        pdf.set_font(
            "Arial",
            'B',
            15
        )

        pdf.cell(
            0,
            10,
            f"{class_name} ({start_date} - {end_date})",
            ln=True,
            align='C'
        )

        pdf.ln(4)

        headers = [
            "Adm No",
            "Student Name",
            "W.D",
            "P",
            "A",
            "L",
            "V.L",
            "Att%"
        ]

        widths = [
            18,
            70,
            12,
            12,
            12,
            12,
            12,
            18
        ]

        pdf.set_font(
            "Arial",
            'B',
            8
        )

        for i, h in enumerate(headers):

            pdf.cell(
                widths[i],
                8,
                h,
                border=1,
                align='C'
            )

        pdf.ln()

        pdf.set_font(
            "Arial",
            '',
            7
        )

        for _, row in df_detail.iterrows():

            values = [
                str(row["Admission No"]),
                str(row["Student Name"])[:40],
                str(row["Working_Days"]),
                str(row["Present"]),
                str(row["Absent"]),
                str(row["Late"]),
                str(row["Very_Late"]),
                f'{row["Attendance %"]:.2f}'
            ]

            # ------------------------------------------------
            # HIGHLIGHT LOGIC
            # ------------------------------------------------

            is_absent = (
                row["Absent"]
                >= absent_highlight_threshold
            )

            is_late = (
                row["Late"]
                >= late_highlight_threshold
                or
                row["Very_Late"]
                >= very_late_highlight_threshold
            )

            if is_absent:

                pdf.set_fill_color(
                    249,
                    73,
                    73
                )

            elif is_late:

                pdf.set_fill_color(
                    255,
                    255,
                    0
                )

            else:

                pdf.set_fill_color(
                    255,
                    255,
                    255
                )

            for i, value in enumerate(values):

                pdf.cell(
                    widths[i],
                    7,
                    value,
                    border=1,
                    align='C',
                    fill=True
                )

            pdf.ln()

    # ========================================================
    # RETURN PDF
    # ========================================================

    pdf_bytes = pdf.output(
        dest='S'
    )

    if isinstance(pdf_bytes, str):

        pdf_bytes = pdf_bytes.encode(
            'latin1'
        )

    return bytes(pdf_bytes)


# ============================================================
# UPLOAD FILE
# ============================================================

uploaded_file = st.file_uploader(
    "Upload your attendance summary Excel file",
    type=["xls", "xlsx"],
    key=f"file_uploader_{st.session_state.file_uploader_key}"
)

if not uploaded_file:

    st.info(
        "Upload your attendance summary file to continue."
    )

    st.stop()


# ============================================================
# READ EXCEL
# ============================================================

@st.cache_data(ttl=600)
def read_excel(
    file,
    sheet_name_hint=""
):

    try:

        if sheet_name_hint:

            return pd.read_excel(
                file,
                sheet_name=sheet_name_hint,
                engine="openpyxl"
            )

        else:

            return pd.read_excel(
                file,
                engine="openpyxl"
            )

    except Exception:

        try:

            return pd.read_excel(
                file,
                engine="openpyxl"
            )

        except Exception as e2:

            st.error(
                f"Error reading file: {e2}"
            )

            return None


# ============================================================
# LOAD DATA
# ============================================================

try:

    df = read_excel(uploaded_file)

    if df is None:

        st.error(
            "Could not read the uploaded Excel file."
        )

        st.stop()

except Exception as e:

    st.error(
        f"Could not read the uploaded Excel file: {e}"
    )

    st.stop()


# ============================================================
# PREVIEW
# ============================================================

st.subheader(
    "Preview of your data"
)

st.dataframe(
    df.head()
)


# ============================================================
# TRANSFORMATION SETTINGS
# ============================================================

st.subheader(
    "Transformation Settings"
)


# ============================================================
# COURSE / CLASS COLUMN
# ============================================================

course_column_candidates = [
    'course_name',
    'Course Name',
    'Class',
    'Grade',
    'Section'
]

course_column = None

for candidate in course_column_candidates:

    if candidate in df.columns:

        course_column = candidate
        break


if course_column:

    st.success(
        f"Detected course/class column: '{course_column}'"
    )

    unique_courses = (
        df[course_column]
        .drop_duplicates()
        .tolist()
    )

    st.write(
        f"Found {len(unique_courses)} unique course/class values:"
    )

    st.write(
        unique_courses
    )

    # --------------------------------------------------------
    # COURSE TO CLASS MAPPING
    # --------------------------------------------------------

    st.subheader(
        "Course to Class Mapping"
    )

    st.write(
        "Please map each course name to a standardized class name:"
    )

    class_mapping = {}

    default_classes = {

        "7th Year": "GRADE 07",
        "6th Year": "GRADE 06",
        "5th Year": "GRADE 05",
        "4th Year": "GRADE 04",
        "3rd Year": "GRADE 03",
        "2nd Year": "GRADE 02",
        "1st Year": "GRADE 01"

    }

    for course in unique_courses:

        if pd.isna(course):

            default_class = "UNASSIGNED"

        else:

            default_class = None

            for key, value in default_classes.items():

                if key in str(course):

                    default_class = value
                    break

        mapped_class = st.text_input(
            f"Map '{course}' to class:",
            value=(
                default_class
                if default_class
                else f"GRADE {course}"
            ),
            key=f"map_{course}"
        )

        class_mapping[course] = (
            mapped_class.strip()
        )

    class_list = list(
        set(class_mapping.values())
    )

else:

    st.warning(
        "Could not detect a course/class column in your data."
    )

    class_names = st.text_area(
        "Enter class names (one per line)",
        value=(
            "GRADE 01\n"
            "GRADE 02\n"
            "GRADE 03\n"
            "GRADE 04\n"
            "GRADE 05\n"
            "GRADE 06\n"
            "GRADE 07"
        ),
        help=(
            "Enter the class names that should appear "
            "in the output. One class per line."
        )
    )

    class_list = [
        name.strip()
        for name in class_names.split('\n')
        if name.strip()
    ]

    class_mapping = {}


# ============================================================
# WORKING DAYS SETTINGS
# ============================================================

st.subheader(
    "Working Days Settings"
)

st.info(
    """
Working days are calculated individually for each student.

The application uses:

    Working Days = Present + Absent

Therefore, if a student joined the institute late, their working
days will automatically be lower than students who were present
for the entire attendance period.

Attendance percentage is calculated using that student's own
working days.
"""
)


# ------------------------------------------------------------
# Informational maximum working days
# ------------------------------------------------------------

auto_working_days = detect_working_days(df)

if auto_working_days:

    st.success(
        f"Maximum working days found in the uploaded data: "
        f"{auto_working_days}"
    )

else:

    st.warning(
        "Could not detect working days from Present + Absent."
    )


# ------------------------------------------------------------
# Manual override
# ------------------------------------------------------------

use_manual = st.checkbox(
    "Use a manual default working-days value"
)

if use_manual:

    working_days = st.number_input(
        "Default working days",
        min_value=1,
        max_value=365,
        value=(
            st.session_state.working_days
            if st.session_state.working_days
            else (
                auto_working_days
                if auto_working_days
                else 1
            )
        ),
        help=(
            "This is only used when a student's working days "
            "cannot be calculated from Present + Absent. "
            "It does not replace each student's own working days."
        )
    )

else:

    working_days = (
        auto_working_days
        if auto_working_days
        else 1
    )


# ============================================================
# HIGHLIGHT SETTINGS
# ============================================================

st.subheader(
    "Late Comer Highlight Settings"
)

late_highlight_threshold = st.number_input(
    "Highlight students if Late days are greater than or equal to:",
    min_value=0,
    max_value=365,
    value=4,
    step=1,
    help=(
        "Late column will be highlighted yellow when "
        "Late days reach this number."
    )
)

very_late_highlight_threshold = st.number_input(
    "Highlight students if Very Late days are greater than or equal to:",
    min_value=0,
    max_value=365,
    value=1,
    step=1,
    help=(
        "Very Late column will be highlighted yellow when "
        "Very Late days reach this number."
    )
)

absent_highlight_threshold = st.number_input(
    "Highlight students if Absent days are greater than or equal to:",
    min_value=0,
    max_value=365,
    value=3,
    step=1,
    help=(
        "Student name and Absent column will be highlighted "
        "red when Absent days reach this number."
    )
)


# ============================================================
# INDIVIDUAL WORKING DAYS OVERRIDE
# ============================================================

override_working_days = st.checkbox(
    "Override working days for specific students",
    help=(
        "Enable this to manually change the working days "
        "of individual students."
    )
)

if override_working_days:

    st.subheader(
        "Set Individual Working Days"
    )

    # --------------------------------------------------------
    # IMPORTANT:
    # Start with each student's OWN working days.
    # NOT the maximum working days.
    # --------------------------------------------------------

    admission_col = find_best_column(
        df,
        'Admission No',
        None
    )

    student_col = find_best_column(
        df,
        'Student Name',
        None
    )

    present_col = find_best_column(
        df,
        'Present',
        None
    )

    absent_col = find_best_column(
        df,
        'Absent',
        None
    )

    if not admission_col or not student_col:

        st.error(
            "Could not find Admission No or Student Name columns."
        )

    else:

        temp_df = df[
            [
                admission_col,
                student_col
            ]
        ].copy()

        temp_df.columns = [
            'Admission No',
            'Student Name'
        ]

        # ----------------------------------------------------
        # Calculate each student's own working days
        # ----------------------------------------------------

        if present_col:

            temp_present = pd.to_numeric(
                df[present_col],
                errors='coerce'
            ).fillna(0)

        else:

            temp_present = pd.Series(
                0,
                index=df.index
            )

        if absent_col:

            temp_absent = pd.to_numeric(
                df[absent_col],
                errors='coerce'
            ).fillna(0)

        else:

            temp_absent = pd.Series(
                0,
                index=df.index
            )

        temp_df['Working_Days'] = (
            temp_present + temp_absent
        )

        edited_df = st.data_editor(
            temp_df,
            use_container_width=True,
            column_config={

                "Working_Days":
                    st.column_config.NumberColumn(
                        "Working Days",
                        min_value=1,
                        max_value=365,
                        step=1
                    )

            },
            key="working_days_editor"
        )

        st.session_state.student_working_days = dict(
            zip(
                edited_df['Admission No'],
                edited_df['Working_Days']
            )
        )


# ============================================================
# INDIVIDUAL LATE / VERY LATE OVERRIDE
# ============================================================

override_late_days = st.checkbox(
    "Override Late / Very Late days for specific students",
    help=(
        "Enable this to manually edit Late and Very Late "
        "days per student."
    )
)

if override_late_days:

    st.subheader(
        "Set Individual Late / Very Late Days"
    )

    admission_col = find_best_column(
        df,
        'Admission No',
        'Admission No'
    )

    student_col = find_best_column(
        df,
        'Student Name',
        'Student Name'
    )

    temp_late_df = df[
        [
            admission_col,
            student_col
        ]
    ].copy()

    temp_late_df.columns = [
        'Admission No',
        'Student Name'
    ]

    # --------------------------------------------------------
    # Late
    # --------------------------------------------------------

    if 'Late' in df.columns:

        temp_late_df['Late'] = pd.to_numeric(
            df['Late'],
            errors='coerce'
        ).fillna(0)

    else:

        temp_late_df['Late'] = 0

    # --------------------------------------------------------
    # Very Late
    # --------------------------------------------------------

    if 'Very_Late' in df.columns:

        temp_late_df['Very_Late'] = pd.to_numeric(
            df['Very_Late'],
            errors='coerce'
        ).fillna(0)

    elif 'Very Late' in df.columns:

        temp_late_df['Very_Late'] = pd.to_numeric(
            df['Very Late'],
            errors='coerce'
        ).fillna(0)

    else:

        temp_late_df['Very_Late'] = 0

    edited_late_df = st.data_editor(
        temp_late_df,
        use_container_width=True,
        column_config={

            "Late":
                st.column_config.NumberColumn(
                    "Late",
                    min_value=0,
                    max_value=365,
                    step=1
                ),

            "Very_Late":
                st.column_config.NumberColumn(
                    "Very Late",
                    min_value=0,
                    max_value=365,
                    step=1
                )

        },
        key="late_days_editor"
    )

    st.session_state.student_late_days = dict(
        zip(
            edited_late_df['Admission No'],
            edited_late_df['Late']
        )
    )

    st.session_state.student_very_late_days = dict(
        zip(
            edited_late_df['Admission No'],
            edited_late_df['Very_Late']
        )
    )


# ============================================================
# INDIVIDUAL ABSENT OVERRIDE
# ============================================================

override_absent_days = st.checkbox(
    "Override Absent days for specific students",
    help=(
        "Enable this to manually change Absent days. "
        "Present days and Attendance % will update automatically."
    )
)

if override_absent_days:

    st.subheader(
        "Set Individual Absent Days"
    )

    st.caption(
        "Set Absent to 0 to mark a student as fully present "
        "for their individual working days."
    )

    admission_col = find_best_column(
        df,
        'Admission No',
        'Admission No'
    )

    student_col = find_best_column(
        df,
        'Student Name',
        'Student Name'
    )

    absent_col = find_best_column(
        df,
        'Absent',
        None
    )

    temp_absent_df = df[
        [
            admission_col,
            student_col
        ]
    ].copy()

    temp_absent_df.columns = [
        'Admission No',
        'Student Name'
    ]

    if absent_col:

        temp_absent_df['Absent'] = pd.to_numeric(
            df[absent_col],
            errors='coerce'
        ).fillna(0).astype(int)

    else:

        temp_absent_df['Absent'] = 0

    edited_absent_df = st.data_editor(
        temp_absent_df,
        use_container_width=True,
        column_config={

            "Absent":
                st.column_config.NumberColumn(
                    "Absent",
                    min_value=0,
                    max_value=365,
                    step=1,
                    help=(
                        "Change this student's absent days. "
                        "Present and Attendance % recalculate automatically."
                    )
                )

        },
        key="absent_days_editor"
    )

    st.session_state.student_absent_days = dict(
        zip(
            edited_absent_df['Admission No'],
            edited_absent_df['Absent']
        )
    )


# ============================================================
# SORT CLASS NAMES
# ============================================================

def sort_class_names(class_names):

    def sort_key(name):

        numbers = re.findall(
            r'\d+',
            name
        )

        grade_num = (
            int(numbers[0])
            if numbers
            else 999
        )

        match = re.search(
            r'-\s*([A-Z])$',
            name
        )

        section = (
            match.group(1)
            if match
            else ''
        )

        return (
            grade_num,
            section
        )

    return sorted(
        class_names,
        key=sort_key
    )


# ============================================================
# CREATE EXCEL
# ============================================================

def to_excel_bytes(
    summary_df,
    detailed_dfs,
    sorted_class_names,
    late_threshold,
    very_late_threshold
):

    wb = Workbook()

    # Remove default sheet
    wb.remove(
        wb.active
    )

    # ========================================================
    # SUMMARY SHEET
    # ========================================================

    ws_summary = wb.create_sheet(
        "Class Summary"
    )

    for r in dataframe_to_rows(
        summary_df,
        index=False,
        header=True
    ):

        ws_summary.append(r)

    start_date, end_date = extract_date_range(
        uploaded_file.name
    )

    summary_title = (
        f"ATTENDANCE SUMMARY - "
        f"{start_date} - {end_date}"
    )

    apply_excel_styling(
        ws_summary,
        summary_title,
        is_summary=True
    )

    # ========================================================
    # CLASS SHEETS
    # ========================================================

    for class_name in sorted_class_names:

        if class_name not in detailed_dfs:
            continue

        sheet_name = (
            class_name[:31]
            if len(class_name) > 31
            else class_name
        )

        ws_class = wb.create_sheet(
            sheet_name
        )

        for r in dataframe_to_rows(
            detailed_dfs[class_name],
            index=False,
            header=True
        ):

            ws_class.append(r)

        student_names = (
            detailed_dfs[class_name]['Student Name'].tolist()
            if 'Student Name'
            in detailed_dfs[class_name].columns
            else []
        )

        apply_excel_styling(
            ws_class,
            class_name,
            is_summary=False,
            student_names=student_names,
            late_threshold=late_highlight_threshold,
            very_late_threshold=very_late_highlight_threshold,
            absent_threshold=absent_highlight_threshold
        )

    # ========================================================
    # SAVE
    # ========================================================

    towrite = BytesIO()

    wb.save(
        towrite
    )

    towrite.seek(0)

    return towrite


# ============================================================
# PROCESS REAL DATA
# ============================================================

def process_real_data(
    df,
    class_list,
    course_column,
    class_mapping,
    working_days
):

    detailed_dfs = {}

    # ========================================================
    # CHECK BATCH ID
    # ========================================================

    if 'batch_id' not in df.columns:

        st.error(
            "The column 'batch_id' is required "
            "to split Grade 02 into sections."
        )

        st.stop()

    # ========================================================
    # REQUIRED COLUMNS
    # ========================================================

    required_columns = [
        'Admission No',
        'Student Name',
        'Present',
        'Absent'
    ]

    available_columns = df.columns.tolist()

    column_mapping = {}

    for req_col in required_columns:

        match = process.extractOne(
            req_col,
            available_columns,
            scorer=fuzz.token_sort_ratio
        )

        if match and match[1] > 60:

            column_mapping[req_col] = match[0]

        else:

            column_mapping[req_col] = req_col

            st.warning(
                f"Could not find a matching column "
                f"for '{req_col}'."
            )

    # ========================================================
    # OPTIONAL COLUMNS
    # ========================================================

    optional_columns = [
        'Late',
        'Very_Late',
        'Very Late'
    ]

    for opt_col in optional_columns:

        match = process.extractOne(
            opt_col,
            available_columns,
            scorer=fuzz.token_sort_ratio
        )

        if match and match[1] > 60:

            column_mapping[opt_col] = match[0]

    # ========================================================
    # RENAME
    # ========================================================

    df = df.rename(
        columns=column_mapping
    )

    # ========================================================
    # OPTIONAL COLUMNS DEFAULT
    # ========================================================

    if 'Late' not in df.columns:

        df['Late'] = 0

    if 'Very_Late' not in df.columns:

        if 'Very Late' in df.columns:

            df['Very_Late'] = df[
                'Very Late'
            ]

        else:

            df['Very_Late'] = 0

    # ========================================================
    # NUMERIC CONVERSION
    # ========================================================

    for col in [
        'Present',
        'Absent',
        'Late',
        'Very_Late'
    ]:

        df[col] = pd.to_numeric(
            df[col],
            errors='coerce'
        ).fillna(0)

    # ========================================================
    # MANUAL LATE OVERRIDES
    # ========================================================

    if (
        st.session_state.student_late_days
    ):

        df['Late'] = df[
            'Admission No'
        ].map(
            st.session_state.student_late_days
        ).fillna(
            df['Late']
        )

    # ========================================================
    # MANUAL VERY LATE OVERRIDES
    # ========================================================

    if (
        st.session_state.student_very_late_days
    ):

        df['Very_Late'] = df[
            'Admission No'
        ].map(
            st.session_state.student_very_late_days
        ).fillna(
            df['Very_Late']
        )

    # ========================================================
    # BASE STUDENT WORKING DAYS
    # ========================================================
    #
    # THIS IS THE IMPORTANT CHANGE.
    #
    # Every student gets their own working days:
    #
    # Working Days = Present + Absent
    #
    # We NO LONGER do:
    #
    # Working Days = maximum Present + Absent
    #
    # ========================================================

    df['Working_Days'] = (
        df['Present']
        + df['Absent']
    )

    # ========================================================
    # MANUAL WORKING DAYS OVERRIDES
    # ========================================================

    if (
        st.session_state.student_working_days
    ):

        working_day_overrides = df[
            'Admission No'
        ].map(
            st.session_state.student_working_days
        )

        df['Working_Days'] = (
            working_day_overrides
            .fillna(df['Working_Days'])
        )

    # ========================================================
    # CLEAN WORKING DAYS
    # ========================================================

    df['Working_Days'] = pd.to_numeric(
        df['Working_Days'],
        errors='coerce'
    ).fillna(0)

    df['Working_Days'] = (
        df['Working_Days']
        .clip(lower=0)
    )

    # ========================================================
    # MANUAL ABSENT OVERRIDES
    # ========================================================

    if (
        st.session_state.student_absent_days
    ):

        absent_overrides = df[
            'Admission No'
        ].map(
            st.session_state.student_absent_days
        )

        df['Absent'] = (
            absent_overrides
            .fillna(df['Absent'])
        )

    # ========================================================
    # CLEAN ABSENT
    # ========================================================

    df['Absent'] = pd.to_numeric(
        df['Absent'],
        errors='coerce'
    ).fillna(0)

    df['Absent'] = (
        df['Absent']
        .clip(lower=0)
    )

    # Absent cannot be greater than
    # the student's own working days.

    df['Absent'] = np.minimum(
        df['Absent'],
        df['Working_Days']
    )

    # ========================================================
    # RECALCULATE PRESENT
    # ========================================================
    #
    # This ensures that:
    #
    # Present + Absent = Working Days
    #
    # for every student.
    #
    # ========================================================

    df['Present'] = (
        df['Working_Days']
        - df['Absent']
    )

    df['Present'] = (
        df['Present']
        .clip(lower=0)
    )

    # ========================================================
    # ATTENDANCE PERCENTAGE
    # ========================================================
    #
    # Every student's percentage uses THEIR OWN
    # working days.
    #
    # ========================================================

    df['Attendance %'] = np.where(

        df['Working_Days'] > 0,

        (
            df['Present']
            / df['Working_Days']
        ) * 100,

        0

    )

    df['Attendance %'] = (
        df['Attendance %']
        .round(2)
    )

    # ========================================================
    # CLASS MAPPING
    # ========================================================

    df['Class'] = df[
        course_column
    ].map(
        class_mapping
    )

    # ========================================================
    # SPLIT GRADE 02 BY BATCH ID
    # ========================================================

    def split_grade_2(row):

        if row['Class'] == 'GRADE 02':

            try:

                section = str(
                    row['batch_id']
                ).split('-')[1]

                return (
                    f"GRADE 02 - {section}"
                )

            except Exception:

                return (
                    "GRADE 02 - UNKNOWN"
                )

        return row['Class']

    df['Class'] = df.apply(
        split_grade_2,
        axis=1
    )

    # ========================================================
    # UPDATE CLASS LIST
    # ========================================================

    updated_class_list = []

    for cls in class_list:

        if cls == 'GRADE 02':

            updated_class_list.extend(
                [
                    'GRADE 02 - A',
                    'GRADE 02 - B'
                ]
            )

        else:

            updated_class_list.append(
                cls
            )

    # Preserve order
    class_list = []

    for cls in updated_class_list:

        if cls not in class_list:

            class_list.append(cls)

    # ========================================================
    # GROUP BY CLASS
    # ========================================================

    for class_name in class_list:

        class_data = df[
            df['Class'] == class_name
        ].copy()

        if class_data.empty:
            continue

        output_columns = [

            'Admission No',
            'Student Name',
            'Working_Days',
            'Present',
            'Absent',
            'Late',
            'Very_Late',
            'Attendance %',
            'Class'

        ]

        output_columns = [
            col
            for col in output_columns
            if col in class_data.columns
        ]

        class_data = class_data[
            output_columns
        ]

        detailed_dfs[
            class_name
        ] = class_data

    return detailed_dfs


# ============================================================
# PROCESS BUTTON
# ============================================================

process_button = st.button(
    "Process Attendance Data"
)


if process_button:

    if working_days is None:

        st.error(
            "Please enter the total number of working days."
        )

        st.stop()

    if working_days <= 0:

        st.error(
            "Please enter a valid number of working days."
        )

        st.stop()

    st.session_state.working_days = (
        working_days
    )

    # --------------------------------------------------------
    # PROCESS
    # --------------------------------------------------------

    detailed_dfs = process_real_data(
        df,
        class_list,
        course_column,
        class_mapping,
        working_days
    )

    if not detailed_dfs:

        st.error(
            "No data was processed. "
            "Please check your input and try again."
        )

        st.stop()

    # ========================================================
    # CREATE SUMMARY
    # ========================================================

    summary_data = []

    sorted_class_names = sort_class_names(
        detailed_dfs.keys()
    )

    for class_name in sorted_class_names:

        df_detail = detailed_dfs[
            class_name
        ]

        summary_data.append({

            "Class":
                class_name,

            "Total_Students":
                len(df_detail),

            "Total_Working_Days":
                round(
                    df_detail[
                        "Working_Days"
                    ].mean(),
                    2
                ),

            "Avg_Present":
                round(
                    df_detail[
                        "Present"
                    ].mean(),
                    2
                ),

            "Avg_Absent":
                round(
                    df_detail[
                        "Absent"
                    ].mean(),
                    2
                ),

            "Avg_Late":
                round(
                    df_detail[
                        "Late"
                    ].mean(),
                    2
                ),

            "Avg_Very_Late":
                round(
                    df_detail[
                        "Very_Late"
                    ].mean(),
                    2
                ),

            "Avg_Attendance_Percentage":
                round(
                    df_detail[
                        "Attendance %"
                    ].mean(),
                    2
                )

        })

    summary_df = pd.DataFrame(
        summary_data
    )

    # ========================================================
    # SAVE SESSION RESULTS
    # ========================================================

    st.session_state.processed = True

    st.session_state.summary_df = (
        summary_df
    )

    st.session_state.detailed_dfs = (
        detailed_dfs
    )

    st.session_state.sorted_class_names = (
        sorted_class_names
    )


# ============================================================
# DISPLAY RESULTS
# ============================================================

if st.session_state.processed:

    st.subheader(
        "Preview of Processed Data"
    )

    tab1, tab2 = st.tabs(
        [
            "Summary",
            "Detailed View"
        ]
    )

    # ========================================================
    # SUMMARY
    # ========================================================

    with tab1:

        st.write(
            "Class Summary"
        )

        st.dataframe(
            st.session_state.summary_df,
            use_container_width=True
        )

    # ========================================================
    # DETAILS
    # ========================================================

    with tab2:

        selected_class = st.selectbox(
            "Select class to view details",
            options=(
                st.session_state
                .sorted_class_names
            )
        )

        st.dataframe(
            st.session_state
            .detailed_dfs[
                selected_class
            ],
            use_container_width=True
        )

    # ========================================================
    # EXCEL DOWNLOAD
    # ========================================================

    excel_bytes = to_excel_bytes(
        st.session_state.summary_df,
        st.session_state.detailed_dfs,
        st.session_state.sorted_class_names,
        late_highlight_threshold,
        very_late_highlight_threshold
    )

    col1, col2 = st.columns(2)

    with col1:

        st.download_button(
            label=(
                "Download Detailed Attendance "
                "Report (Excel)"
            ),
            data=excel_bytes,
            file_name=(
                "detailed_attendance_report.xlsx"
            ),
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )

    # ========================================================
    # PDF DOWNLOAD
    # ========================================================

    with col2:

        pdf_bytes = generate_pdf_report(
            st.session_state.summary_df,
            st.session_state.detailed_dfs,
            st.session_state.sorted_class_names,
            uploaded_file.name
        )

        st.info(
            """
**PDF Export Notice**

The PDF export feature is currently under development.
Some formatting issues may still be present.

For the best experience, the Excel download is recommended.
"""
        )

        st.download_button(
            label=(
                "Download Attendance Report (PDF)"
            ),
            data=pdf_bytes,
            file_name=(
                "attendance_report.pdf"
            ),
            mime="application/pdf",
            help=(
                "PDF export is still under development. "
                "Formatting may not be perfect."
            )
        )

    st.success(
        "Attendance data processed successfully! "
        "Download the files above."
    )

    # ========================================================
    # RESET
    # ========================================================

    if st.button(
        "Add a new file",
        key="reset_button"
    ):

        reset_application()

        st.rerun()


elif not process_button:

    st.info(
        "Click the button above to process your "
        "attendance data based on your settings."
    )


# ============================================================
# INSTRUCTIONS
# ============================================================

st.markdown("---")

st.subheader(
    "Instructions"
)

st.markdown(
    """
The app will create:

- A summary sheet with class statistics
- Separate sheets for each class with detailed student attendance records
- Individual Working Days for every student
- Accurate Attendance % based on each student's own Working Days
- Late and Very Late highlighting
- Absent highlighting
- PDF and Excel reports

### Working Days Calculation

For every student:

**Working Days = Present + Absent**

**Attendance % = Present ÷ Working Days × 100**

This means students who joined the institute later will automatically
have fewer working days than students who were enrolled for the entire
attendance period.
"""
)
